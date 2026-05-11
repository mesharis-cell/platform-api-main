#!/usr/bin/env python3
"""
Bacardi import — XLSX extraction step.

Reads:
    .../redbull-asset-alignment-task/output/bacardi-enrichment-2026-04-20/review.xlsx

Writes:
    /tmp/kadence-bacardi-import/inventory.ndjson    (one JSON object per real row)
    /tmp/kadence-bacardi-import/source-checksum.json (sha256 + counts; consumed by the TS importer)

Filters out:
    - rows with empty asset_name
    - the bogus header-repeat row the cleaning agent left behind
      (detected by `type` not being SERIALIZED or POOLED)

Idempotent. Re-run after every regeneration of review.xlsx.

We use openpyxl rather than the JS exceljs library because exceljs crashes
on the file's embedded drawings during reconcile. openpyxl just reads cell
values and ignores drawings — perfect for our use.
"""

import hashlib
import json
import os
from datetime import datetime, timezone

import openpyxl

SOURCE_DIR = "/home/mshari696/apps/kadence/redbull-asset-alignment-task/output/bacardi-enrichment-2026-04-20"
XLSX_PATH = os.path.join(SOURCE_DIR, "review.xlsx")

OUT_DIR = "/tmp/kadence-bacardi-import"
NDJSON_PATH = os.path.join(OUT_DIR, "inventory.ndjson")
CHECKSUM_PATH = os.path.join(OUT_DIR, "source-checksum.json")

EXPECTED_HEADER = [
    "family_name", "asset_name", "original_name", "brand",
    "image", "image_filename", "category", "qty", "type",
    "fy", "flags", "ai_reasoning", "sheet", "row", "location",
    "deplete_raw", "comments",
]
VALID_TYPES = {"SERIALIZED", "POOLED"}


def sha256_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def main() -> None:
    if not os.path.exists(XLSX_PATH):
        raise SystemExit(f"Source XLSX not found: {XLSX_PATH}")
    os.makedirs(OUT_DIR, exist_ok=True)

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    if "Inventory" not in wb.sheetnames:
        raise SystemExit("Inventory sheet missing")
    if "Excluded" not in wb.sheetnames:
        raise SystemExit("Excluded sheet missing")

    inv = wb["Inventory"]
    rows = inv.iter_rows(values_only=True)
    header = list(next(rows))
    if header != EXPECTED_HEADER:
        raise SystemExit(
            "Header mismatch.\n"
            f"  expected: {EXPECTED_HEADER}\n"
            f"  actual  : {header}"
        )

    written = 0
    skipped_header_junk = 0
    skipped_empty = 0

    with open(NDJSON_PATH, "w", encoding="utf-8") as out:
        for r in rows:
            if r[1] is None or str(r[1]).strip() == "":
                skipped_empty += 1
                continue
            t = r[8]  # type column
            if t not in VALID_TYPES:
                skipped_header_junk += 1
                continue

            obj = {
                "family_name": r[0],
                "asset_name": r[1],
                "original_name": r[2],
                "brand": r[3],
                "image_filename": r[5],
                "category": r[6],
                "qty": int(r[7]) if r[7] is not None else 0,
                "type": t,
                "fy": r[9],
                "sheet": r[12],
                "row": int(r[13]) if r[13] is not None else 0,
                "comments": r[16],
            }
            out.write(json.dumps(obj, ensure_ascii=False) + "\n")
            written += 1

    # Excluded sheet: count data rows, skip header
    excl = wb["Excluded"]
    excluded_count = 0
    for i, r in enumerate(excl.iter_rows(values_only=True)):
        if i == 0:
            continue
        if r[0] is not None:
            excluded_count += 1

    sha = sha256_file(XLSX_PATH)
    size = os.path.getsize(XLSX_PATH)
    checksum = {
        "xlsx_path": XLSX_PATH,
        "xlsx_sha256": sha,
        "xlsx_size_bytes": size,
        "inventory_row_count": written,
        "excluded_row_count": excluded_count,
        "skipped_header_junk": skipped_header_junk,
        "skipped_empty_asset_name": skipped_empty,
        "extracted_at": datetime.now(timezone.utc).isoformat(),
    }
    with open(CHECKSUM_PATH, "w", encoding="utf-8") as f:
        json.dump(checksum, f, indent=2)

    print(f"Inventory rows written         : {written}")
    print(f"Skipped (header-junk row)      : {skipped_header_junk}")
    print(f"Skipped (empty asset_name)     : {skipped_empty}")
    print(f"Excluded sheet rows (untouched): {excluded_count}")
    print(f"XLSX SHA256                    : {sha}")
    print(f"Output NDJSON                  : {NDJSON_PATH}")
    print(f"Output checksum                : {CHECKSUM_PATH}")


if __name__ == "__main__":
    main()
